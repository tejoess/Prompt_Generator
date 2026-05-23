from typing import List, Dict
from openpyxl import load_workbook


def split_combined_prompt_text(text: str) -> Dict[str, str]:
    if not text:
        return {"ai_prompt": "", "system_prompt": ""}

    raw = str(text)
    ai_marker = "AI Prompt:"
    sys_marker = "System Prompt:"

    ai_prompt = ""
    system_prompt = ""

    if ai_marker in raw and sys_marker in raw:
        after_ai = raw.split(ai_marker, 1)[1]
        ai_part, sys_part = after_ai.split(sys_marker, 1)
        ai_prompt = ai_part.strip()
        system_prompt = sys_part.strip()

    return {"ai_prompt": ai_prompt, "system_prompt": system_prompt}


def import_excel_library(file_path: str) -> List[Dict]:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    imported = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue

        placeholder = row[1]
        combined = row[2]

        if not placeholder:
            continue

        parsed = split_combined_prompt_text(combined)

        imported.append(
            {
                "source_placeholder": str(placeholder).strip(),
                "ai_prompt": parsed["ai_prompt"],
                "system_prompt": parsed["system_prompt"],
                "raw_combined_text": str(combined) if combined else "",
            }
        )

    return imported