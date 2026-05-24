import re
import traceback
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session

try:
    from openpyxl import load_workbook as _load_workbook
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

from app.database import get_db
from app.models import PromptRecord
from app.schemas import (
    GeneratePromptRequest,
    SavePromptRequest,
    PromptRecordResponse,
    ApplyGroupingRequest,
)
from app.services.prompt_generator import generate_prompts
from app.services.prompt_validator import validate_generated_prompt

router = APIRouter(prefix="/api", tags=["Prompt Generator"])


@router.post("/generate")
def generate_prompt(payload: GeneratePromptRequest):
    try:
        result = generate_prompts(payload)
        validation = validate_generated_prompt(result, payload.document_types)

        response_payload = {
            "prompt_type": result.get("prompt_type"),
            "placeholder_name": result.get("placeholder_name"),
            "extraction_family": result.get("extraction_family"),
            "output_key": result.get("output_key"),
            "special_tags": result.get("special_tags", []),
            "chunk_count": str(result.get("chunk_count", "")),
            "ai_prompt": result.get("ai_prompt"),
            "system_prompt": result.get("system_prompt"),
            "column_header": result.get("column_header"),
            "filters_logic": result.get("filters_logic"),
            "synonyms_logic": result.get("synonyms_logic"),
            "grouping_logic": result.get("grouping_logic"),
            "final_prompt_text": result.get("final_prompt_text", ""),
            "validation": {
                "status": validation.get("status", "invalid"),
                "issues": validation.get("issues", []),
            },
        }

        return JSONResponse(content=response_payload)

    except Exception as e:
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={"detail": f"Prompt generation failed: {str(e)}"}
        )


@router.post("/save")
def save_prompt(payload: SavePromptRequest, db: Session = Depends(get_db)):
    try:
        record = PromptRecord(
            placeholder_name=payload.placeholder_name,
            prompt_type=payload.prompt_type,
            requirement_text=payload.requirement_text,
            extraction_family=payload.extraction_family,
            output_key=payload.output_key,
            special_tags=",".join(payload.special_tags),
            document_types=",".join(payload.document_types),
            chunk_count=payload.chunk_count,
            ai_prompt=payload.ai_prompt,
            system_prompt=payload.system_prompt,
            column_header=payload.column_header,
            filters_logic=payload.filters_logic,
            synonyms_logic=payload.synonyms_logic,
            grouping_logic=payload.grouping_logic,
            final_prompt_text=payload.final_prompt_text,
            validation_status=payload.validation_status,
            validation_issues="\n".join(payload.validation_issues) if payload.validation_issues else None,
        )

        db.add(record)
        db.commit()
        db.refresh(record)

        return {"message": "Saved successfully", "id": record.id}

    except Exception as e:
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={"detail": f"Save failed: {str(e)}"}
        )


@router.get("/prompts", response_model=list[PromptRecordResponse])
def list_prompts(db: Session = Depends(get_db)):
    try:
        return db.query(PromptRecord).order_by(PromptRecord.id.desc()).all()

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Fetch failed: {str(e)}")


@router.get("/prompts/table-placeholders", response_model=list[PromptRecordResponse])
def list_table_placeholders(db: Session = Depends(get_db)):
    try:
        return (
            db.query(PromptRecord)
            .filter(PromptRecord.placeholder_name.like("%##%"))
            .order_by(PromptRecord.placeholder_name)
            .all()
        )
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Fetch failed: {str(e)}")


@router.post("/prompts/apply-grouping")
def apply_grouping(payload: ApplyGroupingRequest, db: Session = Depends(get_db)):
    try:
        records = db.query(PromptRecord).filter(PromptRecord.id.in_(payload.placeholder_ids)).all()
        if not records:
            raise HTTPException(status_code=404, detail="No records found for given IDs")

        names = [r.placeholder_name for r in records]
        grouping_logic = "@grouping=[" + ",".join(f'"{n}"' for n in names) + "]"

        for record in records:
            record.grouping_logic = grouping_logic
            parts = [grouping_logic]
            if record.column_header:
                parts.append(record.column_header)
            if record.filters_logic:
                parts.append(record.filters_logic)
            if record.synonyms_logic:
                parts.append(record.synonyms_logic)
            record.final_prompt_text = "\n".join(parts)

        db.commit()
        return {"message": f"Grouping applied to {len(records)} record(s)", "grouping_logic": grouping_logic}

    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"detail": f"Apply grouping failed: {str(e)}"})


@router.get("/export-excel")
def export_excel(db: Session = Depends(get_db)):
    """Export all saved prompts in the same format as the source AI Library Excel file."""
    if not _OPENPYXL:
        return JSONResponse(status_code=400, content={"detail": "openpyxl not installed"})
    try:
        import io, json as _json
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        records = db.query(PromptRecord).order_by(PromptRecord.id).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Global AI Prompts"

        # Same headers as the source Excel
        headers = ["placeholder", "chunk_count", "ai_prompt", "doctypes", "system_prompt"]
        ws.append(headers)

        # Style the header row
        header_fill = PatternFill("solid", fgColor="4D6076")
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(1, col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=False)

        for r in records:
            # Format doctypes as JSON array string e.g. ["URS","FS"]
            if r.document_types:
                doc_list = [d.strip() for d in r.document_types.split(",") if d.strip()]
                doctypes_str = _json.dumps(doc_list)
            else:
                doctypes_str = "[]"

            chunk = r.chunk_count or "N/A"

            if r.prompt_type == "table_placeholder":
                # For ## rows: collapse final_prompt_text back to literal \n separators
                ai_col  = (r.final_prompt_text or "").replace("\n", "\\n")
                sys_col = ""
            else:
                ai_col  = r.ai_prompt or ""
                sys_col = r.system_prompt or ""

            ws.append([r.placeholder_name, chunk, ai_col, doctypes_str, sys_col])

        # Auto-size column A (placeholder names)
        max_len = max((len(str(ws.cell(row, 1).value or "")) for row in range(1, ws.max_row + 1)), default=10)
        ws.column_dimensions["A"].width = min(max_len + 4, 60)
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 80
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 80

        # Wrap text in ai_prompt and system_prompt columns
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=prompt_library_export.xlsx"},
        )
    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"detail": f"Export failed: {str(e)}"})


@router.post("/import-excel")
def import_excel(db: Session = Depends(get_db)):
    """Import ## table-placeholder rows from the Excel data file into the DB."""
    if not _OPENPYXL:
        return JSONResponse(status_code=400, content={"detail": "openpyxl is not installed. Run: pip install openpyxl"})

    excel_path = Path(__file__).parent.parent.parent / "data" / "AI_Prompt_Library_2026-05-06.xlsx"
    if not excel_path.exists():
        return JSONResponse(status_code=404, content={"detail": f"Excel file not found: {excel_path}"})

    try:
        import json as _json

        wb = _load_workbook(excel_path)
        ws = wb["Global AI Prompts"]

        imported = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            placeholder = str(row[0] or "").strip()
            if not placeholder or "##" not in placeholder:
                continue

            # Skip duplicates
            if db.query(PromptRecord).filter(PromptRecord.placeholder_name == placeholder).first():
                skipped += 1
                continue

            chunk_count = str(row[1] or "N/A").strip()
            raw_text    = str(row[2] or "").strip()
            doctypes_raw = str(row[3] or "").strip()

            # Parse document types (may be JSON array string like '["IO List"]')
            try:
                doc_list = _json.loads(doctypes_raw)
                doc_types_str = ",".join(doc_list) if isinstance(doc_list, list) else doctypes_raw
            except Exception:
                doc_types_str = doctypes_raw

            # Replace literal \n separators with real newlines
            text = raw_text.replace("\\n", "\n")

            # Extract @grouping=[...] – greedy to handle inner brackets in placeholder names
            grouping_logic = None
            m = re.search(r'@grouping=\[.*"\]', text, re.DOTALL)
            if m:
                raw_grouping = m.group(0)
                # Excel uses real newlines between quoted names; normalise to commas
                grouping_logic = re.sub(r'"\s+"', '","', raw_grouping)

            # Extract <column_header>...</column_header>
            column_header = None
            m = re.search(r'<column_header>.*?</column_header>', text, re.DOTALL)
            if m:
                column_header = m.group(0)

            # Extract <filters>...</filters>
            filters_logic = None
            m = re.search(r'<filters>.*?</filters>', text, re.DOTALL)
            if m:
                filters_logic = m.group(0)

            # Extract <synonyms>...</synonyms>
            synonyms_logic = None
            m = re.search(r'<synonyms>.*?</synonyms>', text, re.DOTALL)
            if m:
                synonyms_logic = m.group(0)

            # Rebuild canonical final_prompt_text
            parts = []
            if grouping_logic:
                parts.append(grouping_logic)
            if column_header:
                parts.append(column_header)
            if filters_logic:
                parts.append(filters_logic)
            if synonyms_logic:
                parts.append(synonyms_logic)
            final_prompt_text = "\n".join(parts)

            record = PromptRecord(
                placeholder_name=placeholder,
                prompt_type="table_placeholder",
                requirement_text=None,
                extraction_family="table_placeholder",
                output_key=None,
                special_tags="@grouping" if grouping_logic else "",
                document_types=doc_types_str,
                chunk_count=chunk_count,
                ai_prompt=None,
                system_prompt=None,
                column_header=column_header,
                filters_logic=filters_logic,
                synonyms_logic=synonyms_logic,
                grouping_logic=grouping_logic,
                final_prompt_text=final_prompt_text,
                validation_status="valid",
                validation_issues=None,
            )
            db.add(record)
            imported += 1

        db.commit()
        return {
            "message": f"Import complete: {imported} imported, {skipped} already existed",
            "imported": imported,
            "skipped": skipped,
        }

    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"detail": f"Import failed: {str(e)}"})


@router.delete("/prompts/{record_id}/grouping")
def remove_grouping(record_id: int, db: Session = Depends(get_db)):
    try:
        record = db.query(PromptRecord).filter(PromptRecord.id == record_id).first()
        if not record:
            raise HTTPException(status_code=404, detail="Record not found")

        record.grouping_logic = None
        parts = []
        if record.column_header:
            parts.append(record.column_header)
        if record.filters_logic:
            parts.append(record.filters_logic)
        if record.synonyms_logic:
            parts.append(record.synonyms_logic)
        record.final_prompt_text = "\n".join(parts)

        db.commit()
        return {"message": "Grouping removed"}

    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"detail": f"Remove grouping failed: {str(e)}"})